import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_EXCEL = ROOT / "STOCK 12-08 para revision sale FINAL.xlsx"
CATALOG = ROOT / "products.json"
ASSETS = (ROOT / "assets").resolve()
DEFAULT_STOCK_EXCEL = ROOT / "STOCK.xlsx"


def as_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


excel = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else DEFAULT_EXCEL
stock_excel = Path(sys.argv[2]).resolve() if len(sys.argv) > 2 else DEFAULT_STOCK_EXCEL
if not excel.is_file():
    raise SystemExit(f"No se encontró el Excel: {excel}")


def product_code(sku, color):
    return f"{as_text(sku)}{as_text(color)}".replace(" ", "")


stock_by_code = {}
if stock_excel.is_file():
    stock_workbook = openpyxl.load_workbook(stock_excel, read_only=True, data_only=True)
    stock_sheet = stock_workbook["STOCK"]
    for row in stock_sheet.iter_rows(min_row=2, max_col=18, values_only=True):
        code = product_code(row[1], row[2])
        if not code:
            continue
        record = stock_by_code.setdefault(code, {"carrito": 0, "local": 0, "rio": 0, "bariloche": 0})
        record["carrito"] += int(row[12] or 0)
        record["local"] += int(row[13] or 0)
        record["rio"] += int(row[16] or 0)
        record["bariloche"] += int(row[17] or 0)
    for record in stock_by_code.values():
        record["total"] = sum(record.values())
else:
    print(f"Advertencia: no se encontró el Excel de stock ({stock_excel}); se mostrará stock 0.")

previous = {}
if CATALOG.is_file():
    with CATALOG.open(encoding="utf-8") as file:
        previous = {item["code"]: item for item in json.load(file)}

workbook = openpyxl.load_workbook(excel, read_only=True, data_only=True)
sheet = workbook["Listado de Precios"]
products = []
seen = set()

for sku, color, code, name, retail, sale_price, discount in sheet.iter_rows(min_row=2, max_col=7, values_only=True):
    code = as_text(code)
    if not code or code in seen:
        continue
    seen.add(code)
    old_image = previous.get(code, {}).get("image")
    image = next((candidate for candidate in (old_image, f"assets/{code}.png") if candidate and (ROOT / candidate).is_file()), None)
    products.append({"code": code, "baseCode": as_text(sku) or code, "color": as_text(color), "name": as_text(name) or code, "retail": retail, "salePrice": sale_price, "discount": discount, "image": image})
    products[-1]["stock"] = stock_by_code.get(code, {"carrito": 0, "local": 0, "rio": 0, "bariloche": 0, "total": 0})

products.sort(key=lambda product: product["code"])
with CATALOG.open("w", encoding="utf-8", newline="\n") as file:
    json.dump(products, file, ensure_ascii=False, indent=2)
    file.write("\n")

used_images = {(ROOT / product["image"]).resolve() for product in products if product["image"]}
if any(image.parent != ASSETS for image in used_images):
    raise RuntimeError("Una imagen referenciada está fuera de assets.")
for image in ASSETS.iterdir():
    if image.is_file() and image.resolve() not in used_images:
        image.unlink()

print(f"Catálogo actualizado: {len(products)} productos; {len(used_images)} imágenes conservadas; stock leído para {len(stock_by_code)} códigos.")
